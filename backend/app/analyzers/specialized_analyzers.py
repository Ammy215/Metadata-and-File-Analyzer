import asyncio
from pathlib import Path
from typing import Dict, Any, List
from app.analyzers.base import BaseAnalyzer


class ExcelAnalyzer(BaseAnalyzer):
    """Deep analysis for modern Excel files (.xlsx via openpyxl).

    Legacy .xls (via the unmaintained xlrd library) is intentionally not
    supported - a .xls upload still gets the universal analyzer set
    (Hash/Entropy/Metadata/Keyword/MIME/YARA), just not this deeper
    spreadsheet-specific inspection.
    """

    async def analyze(self, file_path: Path) -> Dict[str, Any]:
        """Analyze Excel spreadsheets."""
        return await asyncio.to_thread(self._analyze_sync, file_path)

    @staticmethod
    def _analyze_sync(file_path: Path) -> Dict[str, Any]:
        results = {
            "sheets": [],
            "total_rows": 0,
            "total_cols": 0,
            "has_formulas": False,
            "has_macros": False,
            "has_external_links": False,
        }

        wb = None
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, data_only=False)

            results["sheets"] = list(wb.sheetnames)
            results["has_macros"] = wb.vba_controls is not None if hasattr(wb, 'vba_controls') else False

            # Check for formulas and external links
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.data_type == 'f':  # Formula
                            results["has_formulas"] = True
                        if cell.value and isinstance(cell.value, str):
                            if cell.value.startswith('=') and '!' in cell.value:
                                results["has_external_links"] = True

                        results["total_cols"] = max(results["total_cols"], cell.column)
                    results["total_rows"] = max(results["total_rows"], cell.row)
        except Exception:
            pass
        finally:
            if wb is not None:
                wb.close()

        return results


class MarkdownAnalyzer(BaseAnalyzer):
    """Analysis for MARKDOWN files (.md)."""

    async def analyze(self, file_path: Path) -> Dict[str, Any]:
        """Analyze Markdown files."""
        return await asyncio.to_thread(self._analyze_sync, file_path)

    @staticmethod
    def _analyze_sync(file_path: Path) -> Dict[str, Any]:
        results = {
            "headings": [],
            "links": [],
            "has_code_blocks": False,
            "has_html": False,
            "word_count": 0,
        }

        try:
            import re
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                content = f.read()

            # Count headings
            results["headings"] = len(re.findall(r'^#+\s', content, re.MULTILINE))

            # Find links
            links = re.findall(r'\[([^\]]+)\]\(([^)]+)\)', content)
            results["links"] = [link[1] for link in links]

            # Check for code blocks
            results["has_code_blocks"] = '```' in content or '~~~' in content

            # Check for embedded HTML
            results["has_html"] = '<' in content and '>' in content

            # Word count
            results["word_count"] = len(content.split())

        except Exception:
            pass

        return results


def _parse_frame_rate(rate_str: str) -> float:
    """Parse ffprobe's 'N/D' frame-rate string without eval()."""
    try:
        num, _, den = rate_str.partition('/')
        den = den or '1'
        return float(num) / float(den) if float(den) != 0 else 0.0
    except (ValueError, ZeroDivisionError):
        return 0.0


def _run_ffprobe(file_path: Path):
    """Run ffprobe and return parsed JSON, or None with a reason if it
    can't run (binary missing, timeout, bad output) - kept distinct from
    "ran fine, this file has no video/audio streams"."""
    import subprocess
    import json as json_module

    cmd = ['ffprobe', '-v', 'quiet', '-print_format', 'json', '-show_format', '-show_streams', str(file_path)]
    try:
        output = subprocess.run(cmd, capture_output=True, text=True, timeout=10)
    except FileNotFoundError:
        return None, "ffprobe binary not found"
    except subprocess.TimeoutExpired:
        return None, "ffprobe timed out"

    if output.returncode != 0:
        return None, f"ffprobe exited with code {output.returncode}"

    try:
        return json_module.loads(output.stdout), None
    except json_module.JSONDecodeError:
        return None, "ffprobe returned invalid JSON"


class VideoAnalyzer(BaseAnalyzer):
    """Analysis for VIDEO files."""

    async def analyze(self, file_path: Path) -> Dict[str, Any]:
        """Analyze video files."""
        return await asyncio.to_thread(self._analyze_sync, file_path)

    @staticmethod
    def _analyze_sync(file_path: Path) -> Dict[str, Any]:
        results = {
            "duration": 0,
            "width": 0,
            "height": 0,
            "fps": 0,
            "codec": "unknown",
            "bitrate": 0,
            "container": file_path.suffix.lower(),
            "ffprobe_available": True,
        }

        data, error = _run_ffprobe(file_path)
        if data is None:
            results["ffprobe_available"] = False
            return results

        if 'format' in data:
            results["duration"] = float(data['format'].get('duration', 0))
            results["bitrate"] = int(data['format'].get('bit_rate', 0)) / 1000  # kbps

        if 'streams' in data:
            for stream in data['streams']:
                if stream.get('codec_type') == 'video':
                    results["width"] = stream.get('width', 0)
                    results["height"] = stream.get('height', 0)
                    results["fps"] = _parse_frame_rate(stream.get('r_frame_rate', '0/1'))
                    results["codec"] = stream.get('codec_name', 'unknown')

        return results


class AudioAnalyzer(BaseAnalyzer):
    """Analysis for AUDIO files."""

    async def analyze(self, file_path: Path) -> Dict[str, Any]:
        """Analyze audio files."""
        return await asyncio.to_thread(self._analyze_sync, file_path)

    @staticmethod
    def _analyze_sync(file_path: Path) -> Dict[str, Any]:
        results = {
            "duration": 0,
            "sample_rate": 0,
            "channels": 0,
            "codec": "unknown",
            "bitrate": 0,
            "container": file_path.suffix.lower(),
            "ffprobe_available": True,
        }

        data, error = _run_ffprobe(file_path)
        if data is None:
            results["ffprobe_available"] = False
            return results

        if 'format' in data:
            results["duration"] = float(data['format'].get('duration', 0))
            results["bitrate"] = int(data['format'].get('bit_rate', 0)) / 1000  # kbps

        if 'streams' in data:
            for stream in data['streams']:
                if stream.get('codec_type') == 'audio':
                    results["sample_rate"] = stream.get('sample_rate', 0)
                    results["channels"] = stream.get('channels', 0)
                    results["codec"] = stream.get('codec_name', 'unknown')

        return results


class PDFAnalyzer(BaseAnalyzer):
    """Deep analysis for PDF files."""

    # Raw-byte marker scan rather than a full PDF object-graph walk: this
    # catches the common case (markers present as literal bytes) without
    # depending on pdfminer's internal object-resolution API, which is
    # fragile across pdfplumber/pdfminer versions. A PDF that hides these
    # inside a compressed object stream could evade this - same tradeoff
    # most lightweight PDF scanners make; flagged as a heuristic, not a
    # guarantee.
    _JS_MARKERS = (b"/JavaScript", b"/JS")
    _FORM_MARKERS = (b"/AcroForm",)
    _EMBEDDED_MARKERS = (b"/EmbeddedFile", b"/EmbeddedFiles")

    async def analyze(self, file_path: Path) -> Dict[str, Any]:
        """
        Perform deep inspection of PDF.

        Args:
            file_path: Path to PDF file

        Returns:
            Dictionary with PDF analysis results
        """
        return await asyncio.to_thread(self._analyze_sync, file_path)

    @classmethod
    def _analyze_sync(cls, file_path: Path) -> Dict[str, Any]:
        results = {
            "pages": 0,
            "has_embedded_objects": False,
            "has_javascript": False,
            "has_forms": False,
            "text_content": "",
        }

        try:
            import pdfplumber

            with pdfplumber.open(file_path) as pdf:
                results["pages"] = len(pdf.pages)

                # Extract all text
                text_parts = []
                for page in pdf.pages:
                    text = page.extract_text()
                    if text:
                        text_parts.append(text)

                results["text_content"] = "\n".join(text_parts)[:5000]  # Limit to 5000 chars
        except Exception:
            pass

        try:
            with open(file_path, "rb") as fh:
                raw = fh.read()
            results["has_javascript"] = any(marker in raw for marker in cls._JS_MARKERS)
            results["has_forms"] = any(marker in raw for marker in cls._FORM_MARKERS)
            results["has_embedded_objects"] = any(marker in raw for marker in cls._EMBEDDED_MARKERS)
        except Exception:
            pass

        return results


class EXEAnalyzer(BaseAnalyzer):
    """Static analysis for EXE/PE files."""

    # Windows API names commonly used by process-injection/shellcode-loader
    # malware. Deliberately the same list KeywordAnalyzer's CRITICAL tier
    # already flags in binary string-scans - keeping the two lists aligned
    # means this field and the keyword-match findings agree on what counts
    # as suspicious, rather than silently drifting apart.
    SUSPICIOUS_IMPORTS = {
        "CreateRemoteThread", "VirtualAlloc", "VirtualAllocEx", "WriteProcessMemory",
        "WinExec", "ShellExecuteA", "ShellExecuteW", "URLDownloadToFileA",
        "URLDownloadToFileW", "SetWindowsHookExA", "SetWindowsHookExW",
    }

    async def analyze(self, file_path: Path) -> Dict[str, Any]:
        """
        Perform static analysis on executable.

        Args:
            file_path: Path to EXE file

        Returns:
            Dictionary with PE analysis results
        """
        return await asyncio.to_thread(self._analyze_sync, file_path)

    @classmethod
    def _analyze_sync(cls, file_path: Path) -> Dict[str, Any]:
        results = {
            "is_pe": False,
            "imports": [],
            "sections": [],
            "has_suspicious_strings": False,
        }

        pe = None
        try:
            import pefile

            pe = pefile.PE(str(file_path))
            results["is_pe"] = True

            # Extract imports
            if hasattr(pe, 'DIRECTORY_ENTRY_IMPORT'):
                for entry in pe.DIRECTORY_ENTRY_IMPORT:
                    for imp in entry.imports:
                        if imp.name:
                            results["imports"].append(imp.name.decode('utf-8', errors='ignore'))

            results["has_suspicious_strings"] = any(
                name in cls.SUSPICIOUS_IMPORTS for name in results["imports"]
            )

            # Extract sections
            for section in pe.sections:
                results["sections"].append({
                    "name": section.Name.decode('utf-8', errors='ignore').rstrip('\x00'),
                    "size": section.SizeOfRawData,
                    "entropy": section.get_entropy() if hasattr(section, 'get_entropy') else 0,
                })
        except Exception:
            pass
        finally:
            if pe is not None:
                pe.close()

        return results


_yara_rules_cache: Dict[str, Any] = {"rules": None, "mtime": None, "dir": None}


def _get_compiled_yara_rules():
    """Compile (and cache) YARA rules, recompiling only when the rules
    directory's mtime changes. Without this, yara.compile() would run
    from scratch on every single file analyzed - a real cost once YARA
    is wired into every upload rather than never running at all."""
    import yara
    from pathlib import Path as PathlibPath
    from app.config import settings

    yara_dir = PathlibPath(settings.YARA_RULES_DIR)
    if not yara_dir.exists():
        return None

    current_mtime = yara_dir.stat().st_mtime
    if (_yara_rules_cache["rules"] is not None
            and _yara_rules_cache["dir"] == str(yara_dir)
            and _yara_rules_cache["mtime"] == current_mtime):
        return _yara_rules_cache["rules"]

    rule_files = list(yara_dir.glob("*.yar"))
    if not rule_files:
        return None

    rules = yara.compile(filepaths={str(f): str(f) for f in rule_files})
    _yara_rules_cache.update(rules=rules, mtime=current_mtime, dir=str(yara_dir))
    return rules


class YARAAnalyzer(BaseAnalyzer):
    """YARA rule matching analyzer."""

    async def analyze(self, file_path: Path) -> Dict[str, Any]:
        """
        Match YARA rules against file.

        Args:
            file_path: Path to file

        Returns:
            Dictionary with YARA matches. "yara_available" is False when
            YARA couldn't actually be run (not installed, rules failed to
            compile) - distinct from "ran fine, found nothing" so callers
            don't mistake an environment gap for a clean scan result.
        """
        return await asyncio.to_thread(self._analyze_sync, file_path)

    @staticmethod
    def _analyze_sync(file_path: Path) -> Dict[str, Any]:
        results: Dict[str, Any] = {"matches": [], "yara_available": True}

        try:
            rules = _get_compiled_yara_rules()
        except ImportError:
            results["yara_available"] = False
            return results
        except Exception:
            results["yara_available"] = False
            return results

        if rules is None:
            return results  # no rules directory/files configured - not an error

        try:
            matches = rules.match(str(file_path))
            for match in matches:
                results["matches"].append({
                    "rule_name": match.rule,
                    "severity": "HIGH",
                    "tags": list(match.tags) if hasattr(match, 'tags') else [],
                })
        except Exception:
            results["yara_available"] = False

        return results
